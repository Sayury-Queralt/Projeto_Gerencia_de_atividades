#Bibliotecas
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl import Workbook
import re
from datetime import datetime
import mysql.connector


#Lista com os códigos de afetação da planilha de controle // não necessariamente será usado no projeto
Afetacao_list = [["0","Não Afeta Elemento nem Serviços"],
["1.1","Sem Afetação Elemento e Afetação Parcial Serviços"],
["1.2","Afetação Parcial Elemento e sem Afetação Serviços"],
["3","Afetação Parcial Elemento e Serviços"],
["4.1","Afetação Total Elemento e sem Afetação Serviços"],
["4.2","Sem Afetação Elemento e Afetação Total Serviços"],
["5.1","Afetação Parcial Elemento e Total Serviços"],
["5.2","Afetação Total Elemento e Parcial Serviços"],
["6","Afetação Total Elemento e Serviços"]]

# Carregar o arquivo Excel
wb = load_workbook('TP_Previsao_Host_Sayury.xlsx')
ws = wb.active

#Deleta a ultima coluna (VTP PK)
ws.delete_cols(ws.max_column)

#Classe_TP_Excel = []
Classe_TP_DB = []

#Criar a classe TP
class TP:
    def __init__(self, dados):
        self.Data_Criacao = dados[0]
        self.Origem = int(dados[1])
        self.Descricao = dados[2]
        self.Data = dados[3]
        self.Tipo = dados[4]
        self.Status = dados[5]
        self.Executor = dados[7]
        self.Afetacao = dados[8]
        self.Area = dados[9]
        self.ElementosFull = dados [-2]
        self.ElementosAgr = dados[-1]

#Manipulação do formato dos elementos // apresentar uma forma mais resumida
def ElementosFinal(ElementoPlanilha):
    Elemento_Final = {}
    ElementoInput = ElementoPlanilha
    ElementosSplit = [[p[0],p[1],p[2][:3]] for p in (re.split(r"[_-]",elem) for elem in ElementoInput)]
    SortNF = sorted(ElementosSplit, key=lambda x: x[-1])

    for sub in SortNF:
        K = sub[2]
        V = sub[0]
        if K not in Elemento_Final:
            Elemento_Final[K] = []
        Elemento_Final[K].append(V)
    
    return Elemento_Final

#Manipulação e preparação das informações // conversão das informações da planilha do SIGITM para o formato desejado
def Criar_TP_Excel(dados):
    Tipo_TP = dados[4]
    if Tipo_TP == "S":
        Tipo_TP = "PA"
    else:
        Tipo_TP = "PR"
    dados[4]=Tipo_TP

    area = dados[9]
    if area == "Serviços Internet Core-PS":
        area = "O&M"
    else:
        area = "Engenharia"
    dados[9] = area

    Afetacao = dados[8]
    for A in Afetacao_list:
        if A[1] == Afetacao:
            dados[8] = A[0]
     
    Elemento_entrada_def = dados[-1]
    ElementoFinal = ElementosFinal(Elemento_entrada_def)    
    dados.append(ElementoFinal)
    Classe_TP_Excel.append(TP(dados))

def Criar_TP_DB(dados):
    date = dados[3]
    date_convertor = datetime.strptime(date,"%d/%m/%y %H:%M")
    dados[3] = date_convertor

    Tipo_TP = dados[4]
    if Tipo_TP == "S":
        Tipo_TP = "Pré-aprovada"
    else:
        Tipo_TP = "Programada"
    dados[4]=Tipo_TP

    area = dados[9]
    if area == "Serviços Internet Core-PS":
        area = "O&M"
    else:
        area = "Engenharia"
    dados[9] = area

        
    Elemento_entrada_def = dados[-1]
    ElementoFinal = ElementosFinal(Elemento_entrada_def)    
    dados.append(ElementoFinal)
    Classe_TP_DB.append(TP(dados))


#Adiciona a primeira linha em dados
dados = [cell.value for cell in ws[2]]
element = [dados[-1]]
dados[-1] = element

# Processar as demais linhas do Excel
# Iterar sobre as linhas (ignorando o cabeçalho e a primeira linha)
for row in ws.iter_rows(min_row=3, values_only=True):
    if dados[1] != row[1]:
        #Criar_TP_Excel(dados)
        Criar_TP_DB(dados)
        dados = list(row[:-1])
        element = []
        element.append(row[-1])
        element.sort
        dados.append(element)
        
    else:
        element.append(row[-1])
        dados[-1] = element            

#Criar_TP_Excel(dados)
Criar_TP_DB(dados)

def UniqueValue(lista):
    elem = {item for sub in lista for item in sub}
    if len(elem) == 1:
            return elem.pop()
    else:
        return lista

#Função de Busca // usado para testar a saida das TPs
def buscar_tp():
    while True:
        numero = input("\nDigite o número da TP (ou 'sair' para encerrar): ")
        
        if numero.lower() == 'sair':
            print("Encerrando...")
            break
            
        encontrada = False
        for tp in Classe_TP_Excel:
            if str(tp.Origem) == numero:
                print("\n" + "="*50)
                print(f"TP Nº: {tp.Origem}")
                print(f"Data Criação: {tp.Data_Criacao}")
                print(f"Descrição: {tp.Descricao}")
                print(f"Data Prevista: {tp.Data}")
                print(f"Tipo: {tp.Tipo}")
                print(f"Status: {tp.Status}")
                print(f"Executor: {tp.Executor}")
                print(f"Afetação: {tp.Afetacao}")
                print(f"Aberto por: {tp.Area}")
                print(f"Elementos: {tp.ElementosAgr}")
                
                encontrada = True
                break
                
        if not encontrada:
            print(f"\nATENÇÃO: TP {numero} não encontrada!\n")

#buscar_tp()

#Função para gerar o arquivo excel saida com o formato da planilha de controle atual
def Criar_saida_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Teste"
    #cabecalho = list(vars(Classe_TP[0]).keys())
    cabecalho = ["TP","Descrição","Data validação","Validador","Executor","Tipo","Afetação","Status","Férias","Calendário","Aberto por","POP","Elemento"]
    ws.append(cabecalho)
    for tp in Classe_TP_Excel:
        El = tp.ElementosAgr
        
        v = list(El.values())
        K = list(El.keys())
        V = UniqueValue(v)
                
        K_String = str(K)
        V_String = str(V)
        K_String = K_String.replace("'","")
        V_String = V_String.replace("'","")
        ws.append([tp.Origem,tp.Descricao,tp.Data,"",tp.Executor,tp.Tipo,tp.Afetacao,"Em análise","","Pendente",tp.Area,K_String,V_String])
    wb.save("Saida.xlsx")


def Input_DB_GerenciaTP(Classe_TP_DB):
    cursor = Connection_DB.cursor()

    for tp in Classe_TP_DB:
        El = tp.ElementosAgr
        El_Full = tp.ElementosFull
        
        v = list(El.values())
        K = list(El.keys())
        V = UniqueValue(v)
                
        K_String = str(K)
        V_String = str(V)
        K_String = K_String.replace("'","")
        V_String = V_String.replace("'","")
        F_String = str(El_Full)

        sql = f"INSERT INTO ControleTPs (Sequencia, Descricao, Data, Categoria, Afetacao, Area, Elemento, POPs, Executor, Host) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        Input = [tp.Origem,tp.Descricao,tp.Data,tp.Tipo,tp.Afetacao,tp.Area,K_String,V_String,tp.Executor,F_String]
        cursor.execute(sql,Input)
        Connection_DB.commit()

    print(cursor.rowcount, "Registros inseridos")
    cursor.close()
    Connection_DB.close()

#Input_DB_GerenciaTP(Classe_TP_DB)

def Alterar_DB():
    continuar = True
    while(True):
        cursor = Connection_DB.cursor()
        Sequencia = input("Numero da TP que deseja alterar: ")
        opcao = 9
        while(opcao != "1" and opcao != "2" and opcao != "3"):
            opcao = input("O que deseja alterar? (Opções: 1 - Data | 2 - Elemento | 3 - Status)\n")
    
        if opcao == "1":
            nova_data = input("Digite a nova data e a hora no seguinte formato: dd/mm/yy HH:MM\n")
            date_convertor = datetime.strptime(nova_data,"%d/%m/%y %H:%M")
            sql = f"UPDATE ControleTPs SET Data = %s WHERE Sequencia = %s;)"
            Input = [nova_data,Sequencia]

        elif opcao == "2":
            print("Em desenvolvimento")

        elif opcao == "3":
            novo_Status = input("Digite o novo Status da TP: ")
            sql = f"UPDATE ControleTPs SET Status = %s WHERE Sequencia = %s;)"
            Input = [novo_Status,Sequencia]
        
        C = "c"
        while(C != "s" and C != "n"):
            continuar = input("Deseja continuar? (s/n)")

        if C == "s":
            continuar = True
        elif C == "n":
            break
       
    cursor.execute(sql,Input)
    Connection_DB.commit()

Alterar_DB()
