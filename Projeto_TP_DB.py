#Legado Sayury \(^^)/

#Bibliotecas
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl import Workbook
import re
from datetime import datetime
from dotenv import load_dotenv
import os
import mysql.connector

#Carregar o arquivo com as informações do banco de dados
load_dotenv("Infos_DB.env")

#Conexão com o Banco de Dados no servidor (que está junto com o Grafana)
Connection_DB = mysql.connector.connect(
    host = os.getenv("host"),
    port = os.getenv("port"),
    user = os.getenv("user"),   
    password = os.getenv("password"),
    database = os.getenv("database")
)

# Carregar o arquivo Excel
wb = load_workbook('TP_Criação_Host_Sayury.xlsx')
ws = wb.active

# Carregar o arquivo Excel
#wb2 = load_workbook('Editar_Valores_massivo.xlsx')
#ws2 = wb2.active

#Deleta a ultima coluna (VTP PK)
ws.delete_cols(ws.max_column)

Classe_TP_DB = []

#Criar a classe TP
class TP:
    def __init__(self, dados):
        self.Data_Criacao = dados[0]
        self.Origem = int(dados[1])
        self.Descricao = dados[2]
        self.Data = dados[3]
        self.Tipo = dados[4]
        self.Status = dados[5]
        self.StatusData = "Solicitado"
        self.Executor = dados[7]
        self.Afetacao = dados[8]
        self.Area = dados[9]
        self.ElementosFull = dados [-2]
        self.ElementosAgr = dados[-1]

#Manipulação do formato dos elementos // apresentar uma forma mais resumida
def ElementosFinal(ElementoPlanilha):
    Elemento_Final = {}
    ElementoInput = ElementoPlanilha
    ElementosSplit = [[p[0],p[1],p[2][:3]] for p in (re.split(r"[_-]",elem) for elem in ElementoInput)]
    SortNF = sorted(ElementosSplit, key=lambda x: x[-1])

    for sub in SortNF:
        K = sub[2]
        V = sub[0]
        if K not in Elemento_Final:
            Elemento_Final[K] = []
        Elemento_Final[K].append(V)
    
    return Elemento_Final


def Criar_TP_DB(dados):
    date = dados[3]
    date_convertor = datetime.strptime(date,"%d/%m/%y %H:%M")
    dados[3] = date_convertor

    Tipo_TP = dados[4]
    if Tipo_TP == "S":
        Tipo_TP = "Pré-aprovada"
    else:
        Tipo_TP = "Programada"
    dados[4]=Tipo_TP

    area = dados[9]
    if area == "Serviços Internet Core-PS":
        area = "O&M"
    else:
        area = "Engenharia"
    dados[9] = area

        
    Elemento_entrada_def = dados[-1]
    ElementoFinal = ElementosFinal(Elemento_entrada_def)    
    dados.append(ElementoFinal)
    Classe_TP_DB.append(TP(dados))


#Adiciona a primeira linha em dados
dados = [cell.value for cell in ws[2]]
element = [dados[-1]]
dados[-1] = element

# Processar as demais linhas do Excel
# Iterar sobre as linhas (ignorando o cabeçalho e a primeira linha)
for row in ws.iter_rows(min_row=3, values_only=True):
    if dados[1] != row[1]:
        Criar_TP_DB(dados)
        dados = list(row[:-1])
        element = []
        element.append(row[-1])
        element.sort
        dados.append(element)
        
    else:
        element.append(row[-1])
        dados[-1] = element            

Criar_TP_DB(dados)

def UniqueValue(lista):
    elem = {item for sub in lista for item in sub}
    if len(elem) == 1:
            return elem.pop()
    else:
        return lista


def Input_DB_GerenciaTP(Classe_TP_DB):
    cursor = Connection_DB.cursor()

    for tp in Classe_TP_DB:
        El = tp.ElementosAgr
        El_Full = tp.ElementosFull
        
        v = list(El.values())
        K = list(El.keys())
        V = UniqueValue(v)
                
        K_String = str(K)
        V_String = str(V)
        K_String = K_String.replace("'","")
        V_String = V_String.replace("'","")
        F_String = str(El_Full)

        sql = f"INSERT IGNORE INTO ControleTPs (Sequencia, Descricao, Data, Categoria, Afetacao, Area, Elemento, POPs, Executor, Host, StatusData, Status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        Input = [tp.Origem,tp.Descricao,tp.Data,tp.Tipo,tp.Afetacao,tp.Area,K_String,V_String,tp.Executor,F_String,tp.StatusData,"Em análise"]
        cursor.execute(sql,Input)
        Connection_DB.commit()

    print("Registros inseridos")
    cursor.close()
    Connection_DB.close()


def Alterar_DB():
    while(True):
        cursor = Connection_DB.cursor()
        Sequencia = input("Numero da TP que deseja alterar: ")
        opcao = 9
        while(opcao != "1" and opcao != "2" and opcao != "3" and opcao != "4"):
            opcao = input("O que deseja alterar? (Opções: 1 - Data | 2 - Elemento | 3 - Status | 4 - StatusData)\n")
    
        if opcao == "1":
            nova_data = input("Digite a nova data e a hora no seguinte formato: dd/mm/yy HH:MM\n")
            date_convertor = datetime.strptime(nova_data,"%d/%m/%y %H:%M")
            sql = f"UPDATE ControleTPs SET Data = %s WHERE Sequencia = %s;)"
            Input = [nova_data,Sequencia]

        elif opcao == "2":
            print("Em desenvolvimento")

        elif opcao == "3":
            novo_Status = input("Digite o novo Status da TP: ")
            sql = f"UPDATE ControleTPs SET Status = %s WHERE Sequencia = %s;"
            Input = [novo_Status,Sequencia]

        elif opcao == "4":
            novo_StatusData = input("Digite o novo StatusData da TP: ")
            sql = f"UPDATE ControleTPs SET StatusData = %s WHERE Sequencia = %s;"
            Input = [novo_StatusData,Sequencia]
        

        cursor.execute(sql,Input)
        Connection_DB.commit()
        
        C = "c"
        while(C != "s" and C != "n"):
            C = input("Deseja fazer outra alteração?\n(s/n): ")

        if C == "s":
            C = True
        elif C == "n":
            cursor.close()
            Connection_DB.close()
            break

    cursor.close()
    Connection_DB.close()
    
    


while(True):
    acao = input("Deseja incluir novas TPs ou fazer alterações nas existentes? I - Incluir | A - Alterar\n")
    while(acao != "I" and acao != "2" and acao != "A"):
        acao = input("Deseja incluir novas TPs ou fazer alterações nas existentes? I - Incluir | A - Alterar\n")
    
    if acao == "I":
        Input_DB_GerenciaTP(Classe_TP_DB)
    if acao == "A":
        Alterar_DB()
    
    c = "c"
    while(c != "s" and c != "n"):
        c = input("Deseja continuar com mais alguma ação?\n(s/n): ")

    if c == "s":
        c = True
    elif c == "n":
        print("Programa encerrado")
        break