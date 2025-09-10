#Legado Sayury \(^^)/

#Bibliotecas
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl import Workbook
import re

#Lista com os códigos de afetação da planilha de controle // não necessariamente será usado no projeto
Afetacao_list = [["0","Não Afeta Elemento nem Serviços"],
["1.1","Sem Afetação Elemento e Afetação Parcial Serviços"],
["1.2","Afetação Parcial Elemento e sem Afetação Serviços"],
["3","Afetação Parcial Elemento e Serviços"],
["4.1","Afetação Total Elemento e sem Afetação Serviços"],
["4.2","Sem Afetação Elemento e Afetação Total Serviços"],
["5.1","Afetação Parcial Elemento e Total Serviços"],
["5.2","Afetação Total Elemento e Parcial Serviços"],
["6","Afetação Total Elemento e Serviços"]]

# Carregar o arquivo Excel
wb = load_workbook('TP_Criação_Host_Sayury.xlsx')
ws = wb.active

#Deleta a ultima coluna (VTP PK)
ws.delete_cols(ws.max_column)

Classe_TP_Excel = []

#Criar a classe TP
class TP:
    def __init__(self, dados):
        self.Data_Criacao = dados[0]
        self.Origem = int(dados[1])
        self.Descricao = dados[2]
        self.Data = dados[3]
        self.Tipo = dados[4]
        self.Status = dados[5]
        self.Executor = dados[7]
        self.Afetacao = dados[8]
        self.Area = dados[9]
        self.ElementosFull = dados [-2]
        self.ElementosAgr = dados[-1]

#Manipulação do formato dos elementos // apresentar uma forma mais resumida
def ElementosFinal(ElementoPlanilha):
    Elemento_Final = {}
    ElementoInput = ElementoPlanilha
    ElementosSplit = [[p[0],p[1],p[2][:3]] for p in (re.split(r"[_-]",elem) for elem in ElementoInput)]
    SortNF = sorted(ElementosSplit, key=lambda x: x[-1])

    for sub in SortNF:
        K = sub[2]
        V = sub[0]
        if K not in Elemento_Final:
            Elemento_Final[K] = []
        Elemento_Final[K].append(V)
    
    return Elemento_Final

#Manipulação e preparação das informações // conversão das informações da planilha do SIGITM para o formato desejado
def Criar_TP_Excel(dados):
    Tipo_TP = dados[4]
    if Tipo_TP == "S":
        Tipo_TP = "PA"
    else:
        Tipo_TP = "PR"
    dados[4]=Tipo_TP

    area = dados[9]
    if area == "Serviços Internet Core-PS":
        area = "O&M"
    else:
        area = "Engenharia"
    dados[9] = area

    Afetacao = dados[8]
    for A in Afetacao_list:
        if A[1] == Afetacao:
            dados[8] = A[0]
     
    Elemento_entrada_def = dados[-1]
    ElementoFinal = ElementosFinal(Elemento_entrada_def)    
    dados.append(ElementoFinal)
    Classe_TP_Excel.append(TP(dados))

#Adiciona a primeira linha em dados
dados = [cell.value for cell in ws[2]]
element = [dados[-1]]
dados[-1] = element

# Processar as demais linhas do Excel
# Iterar sobre as linhas (ignorando o cabeçalho e a primeira linha)
for row in ws.iter_rows(min_row=3, values_only=True):
    if dados[1] != row[1]:
        Criar_TP_Excel(dados)
        dados = list(row[:-1])
        element = []
        element.append(row[-1])
        element.sort
        dados.append(element)
        
    else:
        element.append(row[-1])
        dados[-1] = element            

Criar_TP_Excel(dados)

def UniqueValue(lista):
    elem = {item for sub in lista for item in sub}
    if len(elem) == 1:
            return elem.pop()
    else:
        return lista



#Função para gerar o arquivo excel saida com o formato da planilha de controle atual
def Criar_saida_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Teste"
    #cabecalho = list(vars(Classe_TP[0]).keys())
    cabecalho = ["TP","Descrição","Data validação","Validador","Executor","Tipo","Afetação","Status","Férias","Calendário","Aberto por","POP","Elemento"]
    ws.append(cabecalho)
    for tp in Classe_TP_Excel:
        El = tp.ElementosAgr
        
        v = list(El.values())
        K = list(El.keys())
        V = UniqueValue(v)
                
        K_String = str(K)
        V_String = str(V)
        K_String = K_String.replace("'","")
        V_String = V_String.replace("'","")
        ws.append([tp.Origem,tp.Descricao,tp.Data,"",tp.Executor,tp.Tipo,tp.Afetacao,"Em análise","","Pendente",tp.Area,K_String,V_String])
    wb.save("Saida.xlsx")

Criar_saida_excel()