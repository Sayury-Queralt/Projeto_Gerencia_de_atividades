from openpyxl import load_workbook
from collections import defaultdict

# Carregar o arquivo Excel
wb = load_workbook('Export_atividades.xlsx')
ws = wb.active

#Deleta a ultima coluna (VTP PK)
ws.delete_cols(ws.max_column)

Classe_TP = []
Conj_POP = []

#Criar a classe TP
class TP:
    def __init__(self, dados):
        self.Data_Criacao = dados[0]
        self.Origem = dados[1]
        self.Descricao = dados[2]
        self.Data = dados[3]
        self.Tipo = dados[4]
        self.Status = dados[5]
        self.Executor = dados[7]
        self.Afetacao = dados[8]
        self.ElementosFull = dados [-2]
        self.ElementosAgr = dados[-1]

#Manipulação do formato dos elementos
def ElementosFinal(ElementoPlanilha):
    Elemento_Final = {}
    ElementoInput = ElementoPlanilha
    ElementosSplit = [elem.split("_") for elem in ElementoInput]
    SortNF = sorted(ElementosSplit, key=lambda x: x[-1])

    for sub in SortNF:
        K = sub[2]
        V = sub[0]
        if K not in Elemento_Final:
            Elemento_Final[K] = []
        Elemento_Final[K].append(V)
    return Elemento_Final

#Manipulação e preparação 
def Criar_TP(dados):
    Tipo_TP = dados[4]
    if Tipo_TP == "S":
        Tipo_TP = "Pré-aprovada"
    else:
        Tipo_TP = "Programada"
    dados[4]=Tipo_TP
    Elemento_entrada_def = dados[-1]
    ElementoFinal = ElementosFinal(Elemento_entrada_def)    
    dados.append(ElementoFinal)
    Classe_TP.append(TP(dados))

#Adiciona a primeira linha em dados
dados = [cell.value for cell in ws[2]]
element = [dados[-1]]
dados[-1] = element

# Processar as demais linhas do Excel
# Iterar sobre as linhas (ignorando o cabeçalho e a primeira linha)
for row in ws.iter_rows(min_row=3, values_only=True):
    if dados[1] != row[1]:
        Criar_TP(dados)
        dados = list(row[:-2])
        element = []
        element.append(row[-1])
        element.sort
        dados.append(element)
        
    else:
        element.append(row[-1])
        dados[-1] = element            
Criar_TP(dados)

for tp in Classe_TP:
    print(f"TP Nº: {tp.Origem}")
    print(f"Elementos: {tp.ElementosFull}")
    print(f"Elementos: {tp.ElementosAgr}")